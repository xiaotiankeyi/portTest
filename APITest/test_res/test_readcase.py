import json
import os

import openpyxl

from APITest.data.data_yaml import add_yaml_data, rand_yaml_data
from APITest.log.log import handle
from APITest.public.requestway import RunMain
from APITest.public.start import startup

Base_path = os.path.dirname(os.path.dirname(__file__))
file_path = os.path.join(Base_path + '/data/Testdata.xlsx')
yaml_path = os.path.join(Base_path + '/data/config_data.yaml')
log_path = os.path.join(Base_path + '/log/outputLog.log')


# logger = handle(file = log_path)


class Testcase(startup):
    logger = handle(file=log_path)
    obj = RunMain()
    file = file_path
    ws = openpyxl.load_workbook(file)
    df = ws.active

    def test_readcase(self):

        for k in range(2, 5):
            self.data_list = {
                '用例编号': self.df.cell(k, 1).value,
                '测试项': self.df.cell(k, 2).value,
                '测试点': self.df.cell(k, 3).value,
                '是否执行': self.df.cell(k, 4).value,
                '数据依赖': self.df.cell(k, 5).value,
                '数据处理方式': self.df.cell(k, 6).value,
                '要处理的值': self.df.cell(k, 7).value,
                'Host': self.df.cell(k, 8).value,
                'URLPath': self.df.cell(k, 9).value,
                'Method': self.df.cell(k, 10).value,
                'RequestDataType': self.df.cell(k, 11).value,
                'RequestData': self.df.cell(k, 12).value,
                'verifyResult': self.df.cell(k, 13).value,
                '断言字段': self.df.cell(k, 14).value,
                'testresult': self.df.cell(k, 15).value,
            }
            self.url = self.data_list['Host'] + self.data_list['URLPath']
            self.method = self.data_list['Method']
            self.headers = eval(self.data_list['RequestDataType'])
            self.data = eval(self.data_list['RequestData'])  # 获取请求参数把字符串转化为字典
            self.data = json.dumps(self.data)  # 在把字典转化为json字符串
            assertField = self.data_list['断言字段']

            if self.data_list['是否执行'] == 'Y':  # 判断用例是否执行
                if self.data_list['数据依赖'] == 'yes':  # 判断是否有数据处理操作
                    # print('执行:', self.data_list)

                    # 判断数据处理方式且值不能为空
                    if self.data_list['数据处理方式'] == "input" and self.data_list['要处理的值'] is not None:
                        # print('input有执行......')
                        # 把上面self.data为json字符串的转化为dict,以便对数据进行处理赋值
                        self.data = eval(self.data)
                        f = eval(self.data_list['获取的值'])  # 获取需要读取的值
                        for val in f:
                            input_data = rand_yaml_data(
                                file=yaml_path, key=val)  # 向yaml文件读取依赖值
                            self.data[val] = input_data  # 对请求数据进行改造
                        self.data = json.dumps(self.data)  # 在把字典转化为json字符串

                    returnResult = self.obj.run_main(
                        self.url, self.method, self.data, self.headers)  # 发起请求
                    # print('返回body类型和数据:', type(returnResult), returnResult)

                    if self.data_list['数据处理方式'] == "output" and self.data_list['要处理的值'] is not None:
                        # print('output有执行.....')

                        for key, y in eval(
                                self.data_list['要处理的值']).items():  # 获取需要处理的值在添加进yaml
                            add_yaml_data(
                                file=yaml_path, key=key, values=returnResult[y])
                    try:
                        if not self.assertEqual(
                                eval(
                                    self.data_list['verifyResult'])[assertField],
                                returnResult[assertField]):  # 断言的值在返回body中
                            print('断言成功')
                            v = 'O%s' % k
                            # print(v)        #写入内容单元格
                            self.df[v] = 'pass'
                            self.logger.debug(
                                '[%s]接口测试成功,返回msg[%s]' %
                                (self.data_list['测试项'], returnResult[assertField]))
                            continue
                    except Exception as e:
                        # print(e)
                        print('断言失败')
                        v = 'O%s' % k
                        # print(v)        #写入内容单元格
                        self.df[v] = 'Fail'
                        self.logger.debug(
                            '测试点[%s]接口测试失败,body返回msg[%s],错误信息%s' %
                            (self.data_list['测试点'], returnResult[assertField], e))
                        continue
                elif self.data_list['数据依赖'] == 'no':  # 没有数据依赖
                    # print('执行', self.data_list)
                    returnResult = self.obj.run_main(
                        self.url, self.method, self.data, self.headers)  # 发起请求
                    # print('返回body类型和数据:', type(returnResult), returnResult)
                    try:
                        if not self.assertEqual(
                                eval(
                                    self.data_list['verifyResult'])[assertField],
                                returnResult[assertField]):  # 断言的值在返回body中
                            print('断言成功')
                            v = 'O%s' % k
                            # print(v)        #写入内容单元格
                            self.df[v] = 'pass'
                            self.logger.debug(
                                '测试点[%s]接口测试成功,body返回msg[%s]' %
                                (self.data_list['测试点'], returnResult[assertField]))
                            continue
                    except Exception as e:
                        # print(e)
                        print('断言失败')
                        v = 'O%s' % k
                        # print(v)        #写入内容单元格
                        self.df[v] = 'Fail'
                        self.logger.debug(
                            '测试点[%s]接口测试失败,body返回msg[%s],错误信息%s' %
                            (self.data_list['测试点'], returnResult[assertField], e))
                        continue
            elif self.data_list['是否执行'] == 'N':
                # print('不执行', data_list)
                continue
                pass
        self.ws.save(self.file)


if __name__ == "__main__":
    a = Testcase()
    a.test_readcase()
    # input_data = rand_yaml_data(file=yaml_path, key='db')  # 向yaml文件读取依赖值
    # print(input_data)
    # logger.debug('测试')

    pass
